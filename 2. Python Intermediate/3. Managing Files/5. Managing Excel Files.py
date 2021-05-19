"""
To read Excel files you need to install an additional module.

Many modules are available, such as:
    . XLRD
    . OpenPyXl
    . XlslWriter
    . PyExcel
    . PyLightXL
    . Pandas
"""

# Imports --------------------------------------------------------------------------------------------------------------
import openpyxl
import os

# File Path ------------------------------------------------------------------------------------------------------------
current_path = os.getcwd()
full_path = os.path.join(current_path, 'Files', 'pipe_schedules.xlsx')

# Create Workbook Object -----------------------------------------------------------------------------------------------
wb_object = openpyxl.load_workbook(full_path, read_only=True, data_only=True)
print(wb_object.sheetnames)

# Sheet Object ---------------------------------------------------------------------------------------------------------
sheet_object = wb_object.active
# sheet_object = wb_object.get_sheet_by_name('sheet_name')

print(sheet_object.max_row)
print(sheet_object.max_column)

# Cell Object ----------------------------------------------------------------------------------------------------------
cell_object = sheet_object.cell(row=1, column=1)
print(cell_object.value)
print(sheet_object['B1'].value)
print(sheet_object['C1'].value)

# Dealing with values --------------------------------------------------------------------------------------------------
headers = []
for i in range(1, sheet_object.max_column+1):
    cell_object = sheet_object.cell(row=1, column=i)
    headers.append(cell_object.value)
print(headers)

index = sheet_object[1]
headers_2 = []
for item in index:
    headers_2.append(item.value)
print(headers_2)

# Finding specific values ----------------------------------------------------------------------------------------------
standard = 'ASME B36.10'
NPS = 0.75
schedule = 80

thickness = 0
for i in range(1, sheet_object.max_row+1):
    values = [sheet_object.cell(row=i, column=j).value for j in range(1, sheet_object.max_column)]
    print(i, values)

    if values[0] == standard and values[1] == NPS and values[4] == schedule:
        thickness = values[8]
        break
print(f'Thickness {standard} - {NPS} / {schedule}: {thickness} mm')

wb_object.close()
